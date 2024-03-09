import openpyxl
import xlsxwriter
from datetime import datetime
import math
from openpyxl.styles import Font
path = 'D:/project xll/Sector Grouping and definition..xlsx'
wb = xlsxwriter.Workbook('grouping.xlsx')
path1 = 'D:/project xll/EmpHist.xlsx'
wrk = openpyxl.load_workbook(filename=path)
sh = wrk.active
wrkbk = openpyxl.load_workbook(filename=path1)
sh1 = wrkbk.active
sheet1 = wb.add_worksheet('Sheet 1')
outwb= xlsxwriter.Workbook('output.xlsx')
sheet_out = outwb.add_worksheet('Sheet out')
date_format=wb.add_format({'num_format':'dd-mm-yyyy'})
k=15
for i in range(1, sh1.max_row+1):
    z=0;str=""
    for j in range(k, 32, 2):
        for p in range(1, sh.max_row+1):
            cell_ob = sh.cell(row=p, column=1)
            cell_ob1 = sh1.cell(row=i, column=j)
            if cell_ob1.value == cell_ob.value and cell_ob1.value !="HAZIRA" :
                sector = sh.cell(row=p, column=2).value
                str=str+" "+sector
                sheet1.write(i-2,z,sector)
                sheet1.write(i-2,z+1,sh1.cell(row=i, column=j+1).value.date(), date_format)
                z=z+2
            elif cell_ob1.value =="HAZIRA" :
               if sh1.cell(row=i, column=j+1).value.date()>=datetime.strptime('2013-05-01 00:00:00','%Y-%m-%d %H:%M:%S').date():
                   sheet1.write(i-2,z,'WS')
                   str=str + " " +"WS"
               if sh1.cell(row=i, column=j+1).value.date()<datetime.strptime('2013-05-01 00:00:00','%Y-%m-%d %H:%M:%S').date():
                   sheet1.write(i-2,z,'MS') 
                   str=str + " " +"MS"
               sheet1.write(i-2,z+1,sh1.cell(row=i, column=j+1).value.date(), date_format)
               z=z+2
               break 
        if(sh1.cell(row=i,column=j+1).value is not None and isinstance((sh1.cell(row=i,column=j+1).value),datetime)): 
            str=str + " " + (sh1.cell(row=i,column=j+1).value.strftime("%Y-%m-%d"))
        a=1
        sheet_out.write(i-1,a-1,sh1.cell(row=i,column=1).value)
        sheet_out.write(i-1,1,sh1.cell(row=i,column=2).value)
    sheet_out.write(i-1,2,str)
wb.close()
path4 = 'D:/python codes/grouping.xlsx'
wrkbook2 = openpyxl.load_workbook(filename=path4)
sh2 = wrkbook2.active 
sheet_out.write(sh2.max_row,a-1,sh1.cell(row=sh1.max_row,column=1).value)
sheet_out.write(sh2.max_row,1,sh1.cell(row=sh1.max_row,column=1).value)
for i in range(1,sh2.max_row+1):
    count=0;sect =0;nec=0;stay_years=0;stay_months=0;stay_m=0;comp=0
    for j in range(1,18,2):
        if sh2.cell(row=i,column=j).value!=sh2.cell(row=i,column=j+2).value and (sh2.cell(row=i,column=j+2).value is not None and sh2.cell(row=i,column=j+1).value is not None): 
            if sh2.cell(row=i,column=j).value=='ES':
                nec+=1
            cell_val = sh2.cell(row=i,column=j+1).value
            sect+=1
            if sh2.cell(row=i,column=j+3).value is not None:
                cell_val1= sh2.cell(row=i,column=j+3).value
                diff = cell_val1-cell_val;years=diff.days//365;months=(diff.days % 365)//30
                if (years > 2 or (years ==2 and months>=10) or sect ==3 and nec<2):
                    count+=1
                if(sh2.cell(row=i,column=j).value=='ES'and sh2.cell(row=i, column=j+1).value.date()>=datetime.strptime('1991-01-01 00:00:00','%Y-%m-%d %H:%M:%S').date() and sh2.cell(row=i, column=j+1).value.date()< datetime.strptime('1992-05-31 00:00:00','%Y-%m-%d %H:%M:%S').date()) and nec<2:
                    count+=1
                    comp=1
                elif(years > 2 or (years ==2 and months>=10)) and nec>=2:
                    count+=2
                    comp=1
                if sh2.cell(row=i,column=j).value=='ES':
                    if(sh2.cell(row=i,column=j+3).value.date()> datetime.strptime('2008-04-01 00:00:00','%Y-%m-%d %H:%M:%S').date() and (years>5 or (years==5 and months >=10)))  :
                        count+=2
                        comp=1
                    stay_years=stay_years+years;stay_m=stay_m+months;stay_years=stay_m//12;temp,temp1 = math.modf(stay_years/12);stay_months=int(12*temp)
        if(sh2.cell(row=i,column=j).value=='ES'and (stay_years > 5 or (stay_years == 5 and stay_months>=10))and (sh2.cell(row=i,column=j+2).value is None and sh2.cell(row=i,column=j+1).value is None )):
            count+=2 
            comp=1
        if sh2.cell(row=i,column=j).value==sh2.cell(row=i,column=j+2).value and (sh2.cell(row=i,column=j+2).value is not None and sh2.cell(row=i,column=j+1).value is not None):
            if sh2.cell(row=i,column=j).value=="ES":
                count+=2
            sh2.cell(row=i,column=j+3).value = sh2.cell(row=i,column=j+1).value
    if(count >=3):
        sheet_out.write(i,3,'YES')
    elif(count<3):
        sheet_out.write(i,3,'NO')
    if(comp>=1):
        sheet_out.write(i,4,'YES')
    elif(comp==0):
        sheet_out.write(i,4,'NO') 
sheet_out.write('C1','POSTING HISTORY')        
sheet_out.write('D1','3 SECTORS COMPLETED')  
sheet_out.write('E1','NE COMPLETED')                     
outwb.close()
paath = 'D:/python codes/output.xlsx'
wrkboook = openpyxl.load_workbook(filename=paath)
shh = wrkboook.active  
cell=shh['A1'];cell1=shh['B1'];cell2=shh['C1'];cell3=shh['D1'];cell4=shh['E1'];cell.font = Font(bold=True);cell1.font = Font(bold=True);cell2.font = Font(bold=True);cell3.font = Font(bold=True);cell4.font = Font(bold=True)
wrkboook.save(filename="D:/project xll/resultf.xlsx")  